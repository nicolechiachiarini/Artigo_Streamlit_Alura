import streamlit as st
import openpyxl
from openpyxl import load_workbook

st.sidebar.title("Site de receitas")
st.sidebar.text("Nesse site de receitas você irá")
st.sidebar.text("encontrar opções de ótimas receitas ")
st.sidebar.text("doces e salgadas. Veja-as a seguir:")
wb = load_workbook(r'C:\Users\Juan\Documents\receitas.xlsx')
sheet = wb.active
opcoes2 = st.sidebar.selectbox(label = "Selecione uma das opções:", options =["Incluir receitas", "Visualizar receitas"])
if (opcoes2 == "Visualizar receitas"):
    opcoes = st.sidebar.selectbox(label = "Selecione uma das opções:", options =["Doces", "Salgados"])
    doces_planilha = []
    salgado_planilha = []
    tipo = 0
    if (opcoes == "Doces"):
        max_linha = sheet.max_row
        for i in range(2, max_linha + 1):
            tipo = sheet.cell(row=i, column=4).value
            if tipo == "doce":
                doces_planilha.append(sheet.cell(row=i, column=1).value)  # guardando em um vetor a coluna produtos da planilha do excel
        doce_titulo = st.sidebar.selectbox(label = "Selecione uma das opções:", options = doces_planilha)
        posicao = 0
        for i in range(2, max_linha + 1):
            linha_doce = sheet.cell(row=i, column=1).value
            if doce_titulo == linha_doce:
                posicao = i
                i = max_linha + 1
        ingredientes = sheet.cell(row=posicao, column=2)
        preparo = sheet.cell(row=posicao, column=3)
        st.title(doce_titulo)
        st.text("Ingredientes:")
        st.write(ingredientes.value)
        st.text("Modo de preparo:")
        st.write(preparo.value)

    if (opcoes == "Salgados"):
        max_linha = sheet.max_row
        for i in range(2, max_linha + 1):
            tipo = sheet.cell(row=i, column=4).value
            if tipo == "salgado":
                salgado_planilha.append(
                    sheet.cell(row=i, column=1).value)  # guardando em um vetor a coluna produtos da planilha do excel
        salgado_titulo = st.sidebar.selectbox(label="Selecione uma das opções:", options=salgado_planilha)
        posicao = 0
        for i in range(2, max_linha + 1):
            linha_salgada = sheet.cell(row=i, column=1).value
            if salgado_titulo == linha_salgada:
                posicao = i
                i = max_linha + 1
        ingredientes = sheet.cell(row=posicao, column=2)
        preparo = sheet.cell(row=posicao, column=3)
        st.title(salgado_titulo)
        st.text("Ingredientes:")
        st.write(ingredientes.value)
        st.text("Modo de preparo:")
        st.write(preparo.value)

elif (opcoes2 == "Incluir receitas"):
    with st.form(key='incluir'):
        tipo = st.selectbox(label="Selecione uma das opções:", options=["doce", "salgado"])
        nome = st.text_area(label="Nome da receita")
        ingredientes = st.text_area(label="Insira os ingredientes")
        preparo = st.text_area(label="Insira o modo de preparo")
        botao = st.form_submit_button('Incluir')  # botao para enviar os dados

    if botao:
        data = (
            (nome, ingredientes, preparo, tipo),
        )  # lista com os valores a serem incluidos

        for i in data:
            sheet.append(i)
        wb.save(r'C:\Users\Juan\Documents\receitas.xlsx')
